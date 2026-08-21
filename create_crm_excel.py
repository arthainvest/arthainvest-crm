import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import os

# Create workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Define colors
header_blue = "1F497D"
header_green = "70AD47"
header_gold = "FFC000"
header_red = "C5504E"
header_navy = "5B9BD5"
header_lime = "92D050"
header_pink = "FF6B6B"
header_purple = "7030A0"
header_dark = "1F4E78"
header_gray = "44546A"
header_darkgray = "203864"

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_header(sheet, row, columns, bg_color, font_color="FFFFFF"):
    """Create a header row with formatting"""
    for col_num, col_name in enumerate(columns, 1):
        cell = sheet.cell(row=row, column=col_num)
        cell.value = col_name
        cell.font = Font(bold=True, color=font_color, size=11)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

def set_column_width(sheet, widths):
    """Set column widths"""
    for col_num, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + col_num)].width = width

# ============================================
# SHEET 1: DASHBOARD
# ============================================
ws_dashboard = wb.create_sheet("Dashboard", 0)

# Title
ws_dashboard.merge_cells("A1:E1")
title_cell = ws_dashboard["A1"]
title_cell.value = "ArthaInvest CRM Dashboard"
title_cell.font = Font(bold=True, size=16, color="FFFFFF")
title_cell.fill = PatternFill(start_color=header_blue, end_color=header_blue, fill_type="solid")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws_dashboard.row_dimensions[1].height = 30

# KPIs
ws_dashboard["A3"] = "📊 KEY METRICS"
ws_dashboard["A3"].font = Font(bold=True, size=12)

kpis = [
    ["Total Clients", "0"],
    ["Active Prospects", "0"],
    ["Open Deals", "0"],
    ["Total Pipeline Value", "₹0"],
    ["Pending Tasks", "0"],
    ["Renewals This Month", "0"],
]

for idx, kpi in enumerate(kpis, 4):
    ws_dashboard[f"A{idx}"] = kpi[0]
    ws_dashboard[f"A{idx}"].font = Font(bold=True)
    ws_dashboard[f"B{idx}"] = kpi[1]
    ws_dashboard[f"B{idx}"].font = Font(size=11)

set_column_width(ws_dashboard, [200, 120, 120, 120, 120])

# ============================================
# SHEET 2: LEADS
# ============================================
ws_leads = wb.create_sheet("Leads")
headers_leads = ["Lead ID", "Name", "Phone", "Email", "Company", "Position",
                 "Source", "Product Interest", "Qualification Status", "Budget",
                 "Timeline", "Created Date", "Last Contact", "Next Follow-up", "Notes"]
create_header(ws_leads, 1, headers_leads, header_blue)

# Add data validation for Source
dv_source = DataValidation(type="list", formula1='"Direct,Referral,LinkedIn,WhatsApp,Email Campaign,Website,Other"')
dv_source.error = 'Please select from the list'
dv_source.errorTitle = 'Invalid Entry'
ws_leads.add_data_validation(dv_source)
dv_source.add(f'G2:G1000')

# Add data validation for Status
dv_status = DataValidation(type="list", formula1='"Warm,Qualified,Not Qualified,On Hold"')
ws_leads.add_data_validation(dv_status)
dv_status.add(f'I2:I1000')

# Format columns
for col in ['J', 'L', 'M', 'N']:
    for row in range(2, 1000):
        ws_leads[f'{col}{row}'].number_format = 'yyyy-mm-dd'

ws_leads['J2'].number_format = '₹#,##0'

set_column_width(ws_leads, [60, 120, 100, 150, 120, 100, 100, 120, 120, 80, 100, 100, 100, 100, 200])

# ============================================
# SHEET 3: CONTACTS
# ============================================
ws_contacts = wb.create_sheet("Contacts")
headers_contacts = ["Contact ID", "Name", "Phone", "Email", "Company", "Position",
                    "Relationship Type", "Related To", "Address", "City", "State",
                    "Preferred Contact Method", "Preferred Time", "Birthday", "Last Contact",
                    "Created Date", "Notes"]
create_header(ws_contacts, 1, headers_contacts, header_green)

dv_relationship = DataValidation(type="list", formula1='"Client,Prospect,Referrer,Influencer,Partner"')
ws_contacts.add_data_validation(dv_relationship)
dv_relationship.add('G2:G1000')

dv_method = DataValidation(type="list", formula1='"WhatsApp,Email,Phone,SMS,In-person"')
ws_contacts.add_data_validation(dv_method)
dv_method.add('L2:L1000')

for col in ['O', 'P', 'Q']:
    for row in range(2, 1000):
        ws_contacts[f'{col}{row}'].number_format = 'yyyy-mm-dd'

set_column_width(ws_contacts, [60, 120, 100, 150, 120, 100, 120, 120, 150, 100, 80, 120, 100, 100, 100, 100, 200])

# ============================================
# SHEET 4: CLIENTS
# ============================================
ws_clients = wb.create_sheet("Clients")
headers_clients = ["Client ID", "Name", "Phone", "Email", "Product", "Folio/Policy No.",
                   "Start Date", "SIP/Premium Amount", "Frequency", "Renewal/Review Date",
                   "Commission Trail", "Last Review", "Status", "Annual Value", "Created Date",
                   "Client Score", "Notes"]
create_header(ws_clients, 1, headers_clients, header_gold, "000000")

dv_status_client = DataValidation(type="list", formula1='"Active,Inactive,Dormant,Churned"')
ws_clients.add_data_validation(dv_status_client)
dv_status_client.add('M2:M1000')

dv_frequency = DataValidation(type="list", formula1='"Monthly,Quarterly,Half-yearly,Annual,One-time"')
ws_clients.add_data_validation(dv_frequency)
dv_frequency.add('I2:I1000')

for col in ['G', 'J', 'L', 'O']:
    for row in range(2, 1000):
        ws_clients[f'{col}{row}'].number_format = 'yyyy-mm-dd'

for col in ['H', 'N']:
    for row in range(2, 1000):
        ws_clients[f'{col}{row}'].number_format = '₹#,##0'

set_column_width(ws_clients, [60, 120, 100, 150, 100, 120, 100, 120, 100, 120, 100, 100, 100, 100, 100, 100, 200])

# ============================================
# SHEET 5: DEALS
# ============================================
ws_deals = wb.create_sheet("Deals")
headers_deals = ["Deal ID", "Deal Name", "Client/Lead Name", "Product", "Expected Value",
                 "Probability %", "Expected Close Date", "Stage", "Owner", "Created Date",
                 "Last Activity", "Key Decision Maker", "Competition", "Notes"]
create_header(ws_deals, 1, headers_deals, header_red)

dv_stage = DataValidation(type="list", formula1='"Prospecting,Qualification,Needs Analysis,Proposal,Negotiation,Closed Won,Closed Lost"')
ws_deals.add_data_validation(dv_stage)
dv_stage.add('H2:H1000')

for col in ['G', 'J', 'K']:
    for row in range(2, 1000):
        ws_deals[f'{col}{row}'].number_format = 'yyyy-mm-dd'

for row in range(2, 1000):
    ws_deals[f'E{row}'].number_format = '₹#,##0'
    ws_deals[f'F{row}'].number_format = '0"%"'

set_column_width(ws_deals, [60, 120, 120, 100, 120, 80, 120, 120, 100, 100, 100, 120, 100, 200])

# ============================================
# SHEET 6: PRODUCTS
# ============================================
ws_products = wb.create_sheet("Products")
headers_products = ["Product ID", "Product Name", "Category", "Provider", "Commission %",
                    "Min Premium", "Max Premium", "Description", "Key Features", "Active", "Created Date"]
create_header(ws_products, 1, headers_products, header_navy)

# Add sample products
sample_products = [
    ["P001", "Tata AIG Term Plan", "Life Insurance", "Tata AIG", 8, 50000, 10000000, "Term life insurance plan", "High coverage, low premium", "Yes", datetime.now()],
    ["P002", "Niva Bupa Health Insurance", "Health Insurance", "Niva Bupa", 12, 10000, 500000, "Comprehensive health coverage", "Cashless treatment, wide network", "Yes", datetime.now()],
    ["P003", "POSP Pension Plan", "Pension", "Government", 5, 100000, 5000000, "Pension scheme", "Tax benefits, retirement planning", "Yes", datetime.now()],
    ["P004", "DSA Investment Plan", "Investment", "Various", 6, 50000, 2000000, "Direct Selling Agent commission", "Flexible, good returns", "Yes", datetime.now()],
]

for row_idx, product in enumerate(sample_products, 2):
    for col_idx, value in enumerate(product, 1):
        ws_products.cell(row=row_idx, column=col_idx, value=value)

dv_active = DataValidation(type="list", formula1='"Yes,No"')
ws_products.add_data_validation(dv_active)
dv_active.add('J2:J1000')

for row in range(2, 1000):
    ws_products[f'E{row}'].number_format = '0"%"'
    ws_products[f'F{row}'].number_format = '₹#,##0'
    ws_products[f'G{row}'].number_format = '₹#,##0'
    ws_products[f'K{row}'].number_format = 'yyyy-mm-dd'

set_column_width(ws_products, [80, 150, 120, 120, 100, 100, 100, 200, 200, 80, 100])

# ============================================
# SHEET 7: TASKS
# ============================================
ws_tasks = wb.create_sheet("Tasks")
headers_tasks = ["Task ID", "Task Description", "Assigned To", "Related To", "Related Type",
                 "Status", "Priority", "Due Date", "Created Date", "Completed Date", "Notes"]
create_header(ws_tasks, 1, headers_tasks, header_lime, "000000")

dv_task_status = DataValidation(type="list", formula1='"Pending,In Progress,Completed,On Hold"')
ws_tasks.add_data_validation(dv_task_status)
dv_task_status.add('F2:F1000')

dv_priority = DataValidation(type="list", formula1='"High,Medium,Low"')
ws_tasks.add_data_validation(dv_priority)
dv_priority.add('G2:G1000')

dv_type = DataValidation(type="list", formula1='"Lead,Client,Deal,Other"')
ws_tasks.add_data_validation(dv_type)
dv_type.add('E2:E1000')

for col in ['H', 'I', 'J']:
    for row in range(2, 1000):
        ws_tasks[f'{col}{row}'].number_format = 'yyyy-mm-dd'

set_column_width(ws_tasks, [60, 200, 100, 120, 100, 100, 100, 100, 100, 100, 200])

# ============================================
# SHEET 8: COMMUNICATIONS
# ============================================
ws_communications = wb.create_sheet("Communications")
headers_communications = ["Communication ID", "Date", "Contact Name", "Channel", "Type", "Message Subject",
                          "Status", "Outcome", "Next Follow-up", "Duration (min)", "Notes"]
create_header(ws_communications, 1, headers_communications, header_pink)

dv_channel = DataValidation(type="list", formula1='"WhatsApp,Email,Phone Call,In-person Meeting,Video Call,SMS"')
ws_communications.add_data_validation(dv_channel)
dv_channel.add('D2:D1000')

dv_comm_type = DataValidation(type="list", formula1='"Inquiry,Follow-up,Proposal,Negotiation,Feedback,Support"')
ws_communications.add_data_validation(dv_comm_type)
dv_comm_type.add('E2:E1000')

dv_comm_status = DataValidation(type="list", formula1='"Sent,Delivered,Opened,Replied,Scheduled,Failed"')
ws_communications.add_data_validation(dv_comm_status)
dv_comm_status.add('G2:G1000')

for row in range(2, 1000):
    ws_communications[f'B{row}'].number_format = 'yyyy-mm-dd hh:mm:ss'
    ws_communications[f'I{row}'].number_format = 'yyyy-mm-dd'

set_column_width(ws_communications, [80, 100, 120, 100, 100, 200, 100, 150, 100, 100, 200])

# ============================================
# SHEET 9: DOCUMENTS
# ============================================
ws_documents = wb.create_sheet("Documents")
headers_documents = ["Document ID", "Document Name", "Type", "Related To", "Client/Lead Name",
                     "Upload Date", "Status", "Expiry Date", "File Link", "Notes"]
create_header(ws_documents, 1, headers_documents, header_purple)

dv_doc_type = DataValidation(type="list", formula1='"Policy Document,KYC Form,Quotation,Proposal,Invoice,Agreement,Medical Report,Other"')
ws_documents.add_data_validation(dv_doc_type)
dv_doc_type.add('C2:C1000')

dv_doc_status = DataValidation(type="list", formula1='"Uploaded,Pending,Expired,Archived"')
ws_documents.add_data_validation(dv_doc_status)
dv_doc_status.add('G2:G1000')

for col in ['F', 'H']:
    for row in range(2, 1000):
        ws_documents[f'{col}{row}'].number_format = 'yyyy-mm-dd'

set_column_width(ws_documents, [80, 150, 120, 120, 120, 100, 100, 100, 300, 200])

# ============================================
# SHEET 10: COMMISSIONS
# ============================================
ws_commissions = wb.create_sheet("Commissions")
headers_commissions = ["Commission ID", "Date", "Agent/DSA", "Policy/Deal No.", "Client Name", "Product",
                       "Commission Amount", "Commission %", "Commission Type", "Status", "Payment Date", "Notes"]
create_header(ws_commissions, 1, headers_commissions, header_dark)

dv_comm_type_comm = DataValidation(type="list", formula1='"Initial,Trail,Bonus,Incentive"')
ws_commissions.add_data_validation(dv_comm_type_comm)
dv_comm_type_comm.add('I2:I1000')

dv_comm_status_comm = DataValidation(type="list", formula1='"Earned,Pending,Paid,Disputed"')
ws_commissions.add_data_validation(dv_comm_status_comm)
dv_comm_status_comm.add('J2:J1000')

for row in range(2, 1000):
    ws_commissions[f'G{row}'].number_format = '₹#,##0.00'
    ws_commissions[f'H{row}'].number_format = '0.00"%"'
    ws_commissions[f'B{row}'].number_format = 'yyyy-mm-dd'
    ws_commissions[f'K{row}'].number_format = 'yyyy-mm-dd'

set_column_width(ws_commissions, [80, 100, 120, 120, 120, 120, 120, 100, 120, 100, 100, 200])

# ============================================
# SHEET 11: REPORTS
# ============================================
ws_reports = wb.create_sheet("Reports")
ws_reports.merge_cells("A1:D1")
report_title = ws_reports["A1"]
report_title.value = "ArthaInvest CRM Reports"
report_title.font = Font(bold=True, size=14, color="FFFFFF")
report_title.fill = PatternFill(start_color=header_gray, end_color=header_gray, fill_type="solid")
report_title.alignment = Alignment(horizontal="center", vertical="center")

ws_reports["A3"] = "MONTHLY SUMMARY"
ws_reports["A3"].font = Font(bold=True, size=12)

summary_headers = ["Metric", "This Month", "Last Month", "Growth %"]
create_header(ws_reports, 4, summary_headers, "E8E8E8")

summary_data = [
    ["New Clients", 0, 0, ""],
    ["New Leads", 0, 0, ""],
    ["Deals Closed", 0, 0, ""],
    ["Total Commission", 0, 0, ""],
    ["Tasks Completed", 0, 0, ""],
    ["Follow-ups Made", 0, 0, ""],
    ["Client Retention Rate %", 0, 0, ""]
]

for row_idx, row_data in enumerate(summary_data, 5):
    for col_idx, value in enumerate(row_data, 1):
        ws_reports.cell(row=row_idx, column=col_idx, value=value)

set_column_width(ws_reports, [150, 100, 100, 100])

# ============================================
# SHEET 12: SETTINGS
# ============================================
ws_settings = wb.create_sheet("Settings")
ws_settings.merge_cells("A1:B1")
settings_title = ws_settings["A1"]
settings_title.value = "CRM SETTINGS"
settings_title.font = Font(bold=True, size=14, color="FFFFFF")
settings_title.fill = PatternFill(start_color=header_darkgray, end_color=header_darkgray, fill_type="solid")
settings_title.alignment = Alignment(horizontal="center", vertical="center")

# Team Settings
ws_settings["A3"] = "TEAM MEMBERS"
ws_settings["A3"].font = Font(bold=True, size=12)

team_headers = ["Name", "Role", "Email", "Phone"]
create_header(ws_settings, 4, team_headers, "E8E8E8")

team_data = [["You", "Owner/Agent", "neemailbox555@gmail.com", ""]]
for row_idx, row_data in enumerate(team_data, 5):
    for col_idx, value in enumerate(row_data, 1):
        ws_settings.cell(row=row_idx, column=col_idx, value=value)

# Commission Rates
ws_settings["A8"] = "COMMISSION RATES (%)"
ws_settings["A8"].font = Font(bold=True, size=12)

commission_headers = ["Product", "Commission Rate"]
create_header(ws_settings, 9, commission_headers, "E8E8E8")

commission_data = [
    ["Tata AIG Term", "8%"],
    ["Niva Bupa Health", "12%"],
    ["POSP Pension", "5%"],
    ["DSA Investment", "6%"],
    ["Other Products", "5%"],
    ["Referral Bonus", "3%"]
]

for row_idx, row_data in enumerate(commission_data, 10):
    for col_idx, value in enumerate(row_data, 1):
        ws_settings.cell(row=row_idx, column=col_idx, value=value)

set_column_width(ws_settings, [150, 150, 150, 150])

# Save the workbook
output_path = r"C:\Users\artha\OneDrive\Desktop\ArthaInvest\ArthaInvest_CRM_Complete.xlsx"
wb.save(output_path)
print("SUCCESS: Excel CRM created successfully!")
print(f"Location: {output_path}")
print("Sheets created: 12 modules")
print("[DONE] Dashboard")
print("[DONE] Leads")
print("[DONE] Contacts")
print("[DONE] Clients")
print("[DONE] Deals")
print("[DONE] Products (with sample data)")
print("[DONE] Tasks")
print("[DONE] Communications")
print("[DONE] Documents")
print("[DONE] Commissions")
print("[DONE] Reports")
print("[DONE] Settings (with commission rates)")
