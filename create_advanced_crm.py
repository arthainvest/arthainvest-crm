import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# Create workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Define colors - Premium theme
COLOR_PRIMARY = "1F4788"
COLOR_LEADS = "4472C4"
COLOR_CONTACTS = "70AD47"
COLOR_CLIENTS = "FFC000"
COLOR_DEALS = "C5504E"
COLOR_PRODUCTS = "5B9BD5"
COLOR_TASKS = "92D050"
COLOR_COMM = "FF6B6B"
COLOR_DOCS = "7030A0"
COLOR_COMM_TRACK = "203864"
COLOR_REPORTS = "44546A"
COLOR_WORKFLOW = "2E75B6"
COLOR_SCORING = "F79646"
COLOR_SEQUENCES = "A6A6A6"

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_header(sheet, row, columns, bg_color, font_color="FFFFFF"):
    for col_num, col_name in enumerate(columns, 1):
        cell = sheet.cell(row=row, column=col_num)
        cell.value = col_name
        cell.font = Font(bold=True, color=font_color, size=11)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

def set_column_width(sheet, widths):
    for col_num, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(col_num)].width = width

# ============================================
# SHEET 1: MASTER DASHBOARD
# ============================================
ws = wb.create_sheet("Dashboard", 0)
ws.merge_cells("A1:H1")
ws["A1"].value = "ArthaInvest CRM - Complete System"
ws["A1"].font = Font(bold=True, size=18, color="FFFFFF")
ws["A1"].fill = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type="solid")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 35

# KPIs Section
ws["A3"] = "REAL-TIME DASHBOARD"
ws["A3"].font = Font(bold=True, size=12, color="FFFFFF")
ws["A3"].fill = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type="solid")

kpis = [
    ["Total Clients", "27", "Active relationships"],
    ["Active Prospects", "12", "In pipeline"],
    ["Open Deals", "0", "Expected revenue"],
    ["Pipeline Value", "0", "Total value"],
    ["Hot Leads", "0", "Score > 80"],
    ["Pending Tasks", "0", "To do"],
    ["This Month Commission", "0", "Earned"],
    ["Client Retention", "100%", "Active clients"],
]

for idx, kpi in enumerate(kpis, 4):
    ws[f"A{idx}"] = kpi[0]
    ws[f"A{idx}"].font = Font(bold=True, size=10)
    ws[f"B{idx}"] = kpi[1]
    ws[f"B{idx}"].font = Font(size=11, bold=True, color=COLOR_PRIMARY)
    ws[f"C{idx}"] = kpi[2]
    ws[f"C{idx}"].font = Font(italic=True, size=9, color="666666")

# Business Info
ws["E3"] = "BUSINESS PROFILE"
ws["E3"].font = Font(bold=True, size=12, color="FFFFFF")
ws["E3"].fill = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type="solid")

business_info = [
    ["Agent Name", "ArthaInvest Agent"],
    ["ARN/Registration", "ARN-267891"],
    ["License Types", "POSP | TATA Bupa | Niva Bupa | DSA"],
    ["Capacity/Month", "8 actionable leads"],
    ["Location", "India"],
    ["Contact", "neemailbox555@gmail.com"],
    ["Last Updated", datetime.now().strftime("%Y-%m-%d")],
]

for idx, info in enumerate(business_info, 4):
    ws[f"E{idx}"] = info[0]
    ws[f"E{idx}"].font = Font(bold=True)
    ws[f"F{idx}"] = info[1]

set_column_width(ws, [200, 100, 200, 50, 200, 250])

# ============================================
# SHEET 2: LEADS (Advanced)
# ============================================
ws_leads = wb.create_sheet("Leads")
headers_leads = ["Lead ID", "Name", "Phone", "Email", "Company", "Position", "Source",
                 "Product Interest", "Qualification Status", "Lead Score", "Budget", "Timeline",
                 "Created Date", "Last Contact", "Next Follow-up", "Lead Source Channel",
                 "Industry", "Company Size", "Decision Timeline", "Notes"]
create_header(ws_leads, 1, headers_leads, COLOR_LEADS)

# Data validation
dv_source = DataValidation(type="list", formula1='"Direct,Referral,LinkedIn,WhatsApp,Email Campaign,Website,Partner,Event"')
ws_leads.add_data_validation(dv_source)
dv_source.add('G2:G1000')

dv_qualification = DataValidation(type="list", formula1='"New,Contacted,Qualified,Proposal Sent,Negotiation,Unqualified,On Hold"')
ws_leads.add_data_validation(dv_qualification)
dv_qualification.add('I2:I1000')

dv_product = DataValidation(type="list", formula1='"Term Insurance,Health Insurance,Pension Plan,Investment,Endowment,Mixed"')
ws_leads.add_data_validation(dv_product)
dv_product.add('H2:H1000')

# Add sample leads
sample_leads = [
    ["L001", "Rajesh Patel", "98765-43210", "rajesh@tech.com", "Tech Solutions", "MD", "LinkedIn", "Term Insurance", "Qualified", "85", "50,00,000", "2 months", "2026-08-01", "2026-08-20", "2026-08-25", "LinkedIn", "Technology", "50-100", "Q3", "Very interested, budget approved"],
    ["L002", "Priya Sharma", "97654-32109", "priya@finance.com", "Finance Corp", "CFO", "Referral", "Health Insurance", "Qualified", "75", "30,00,000", "1 month", "2026-08-05", "2026-08-18", "2026-08-22", "Referral", "Finance", "100-500", "Q2", "Good fit, needs board approval"],
]

for row_idx, lead in enumerate(sample_leads, 2):
    for col_idx, value in enumerate(lead, 1):
        ws_leads.cell(row=row_idx, column=col_idx, value=value)

set_column_width(ws_leads, [60, 120, 100, 150, 120, 100, 100, 120, 120, 80, 100, 80, 100, 100, 100, 150, 100, 100, 100, 200])

# ============================================
# SHEET 3: CONTACTS (Advanced)
# ============================================
ws_contacts = wb.create_sheet("Contacts")
headers_contacts = ["Contact ID", "Name", "Phone", "Email", "Company", "Position",
                    "Relationship Type", "Related To", "Address", "City", "State", "Country",
                    "Preferred Contact Method", "Preferred Time", "Birthday", "LinkedIn", "Notes",
                    "Last Contact", "Contact Frequency", "Risk Level"]
create_header(ws_contacts, 1, headers_contacts, COLOR_CONTACTS)

dv_relationship = DataValidation(type="list", formula1='"Client,Prospect,Referrer,Influencer,Partner,Decision Maker"')
ws_contacts.add_data_validation(dv_relationship)
dv_relationship.add('G2:G1000')

dv_method = DataValidation(type="list", formula1='"WhatsApp,Email,Phone,SMS,In-person,Video Call"')
ws_contacts.add_data_validation(dv_method)
dv_method.add('M2:M1000')

set_column_width(ws_contacts, [60, 120, 100, 150, 120, 100, 120, 120, 150, 100, 80, 100, 120, 100, 100, 150, 200, 100, 100, 100])

# ============================================
# SHEET 4: CLIENTS (Advanced)
# ============================================
ws_clients = wb.create_sheet("Clients")
headers_clients = ["Client ID", "Name", "Phone", "Email", "Product", "Folio/Policy No.",
                   "Provider", "Start Date", "Premium Amount", "Frequency", "Renewal Date",
                   "Commission (%)", "Trail Commission", "Total Premium", "Status",
                   "Annual Value", "Satisfaction Score", "Last Review", "Client Tier",
                   "Notes"]
create_header(ws_clients, 1, headers_clients, COLOR_CLIENTS, "000000")

dv_status = DataValidation(type="list", formula1='"Active,Inactive,Dormant,Churned,Pending Renewal"')
ws_clients.add_data_validation(dv_status)
dv_status.add('O2:O1000')

dv_frequency = DataValidation(type="list", formula1='"Monthly,Quarterly,Half-yearly,Annual,One-time"')
ws_clients.add_data_validation(dv_frequency)
dv_frequency.add('J2:J1000')

dv_tier = DataValidation(type="list", formula1='"Gold,Silver,Bronze,Prospect"')
ws_clients.add_data_validation(dv_tier)
dv_tier.add('S2:S1000')

# Sample clients
sample_clients = [
    ["C001", "Priya Sharma", "98765-43211", "priya@email.com", "Tata AIG Term", "POL-2024-001",
     "Tata AIG", "2024-03-15", "5000", "Monthly", "2026-09-15", "8%", "500", "60000", "Active",
     "60000", "90", "2026-08-15", "Gold", "High-value client"],
    ["C002", "Amit Kumar", "97654-32112", "amit@corp.com", "Niva Bupa Health", "POL-2024-002",
     "Niva Bupa", "2024-06-20", "8000", "Monthly", "2026-12-20", "12%", "800", "96000", "Active",
     "96000", "85", "2026-08-10", "Gold", "Corporate client"],
]

for row_idx, client in enumerate(sample_clients, 2):
    for col_idx, value in enumerate(client, 1):
        ws_clients.cell(row=row_idx, column=col_idx, value=value)

set_column_width(ws_clients, [60, 120, 100, 150, 100, 120, 120, 100, 120, 100, 120, 80, 120, 100, 100, 100, 100, 100, 100, 200])

# ============================================
# SHEET 5: DEALS (Advanced)
# ============================================
ws_deals = wb.create_sheet("Deals")
headers_deals = ["Deal ID", "Deal Name", "Client/Lead Name", "Product", "Expected Value",
                 "Probability %", "Expected Close Date", "Stage", "Owner", "Created Date",
                 "Last Activity", "Decision Maker", "Competition", "Deal Type",
                 "Expected Renewal Value", "Commission Expected", "Notes"]
create_header(ws_deals, 1, headers_deals, COLOR_DEALS)

dv_stage = DataValidation(type="list", formula1='"Lead,Prospect,Qualified,Proposal,Negotiation,Won,Lost"')
ws_deals.add_data_validation(dv_stage)
dv_stage.add('H2:H1000')

dv_deal_type = DataValidation(type="list", formula1='"New Business,Renewal,Cross-sell,Upsell"')
ws_deals.add_data_validation(dv_deal_type)
dv_deal_type.add('N2:N1000')

set_column_width(ws_deals, [60, 120, 120, 100, 120, 80, 120, 120, 100, 100, 100, 120, 100, 120, 150, 150, 200])

# ============================================
# SHEET 6: PRODUCTS (Pre-loaded)
# ============================================
ws_products = wb.create_sheet("Products")
headers_products = ["Product ID", "Product Name", "Category", "Provider", "Commission %",
                    "Min Premium", "Max Premium", "Description", "Key Features",
                    "Active", "Policy Term Options", "Coverage Limit"]
create_header(ws_products, 1, headers_products, COLOR_PRODUCTS)

products_data = [
    ["P001", "Tata AIG Term Plan", "Life Insurance", "Tata AIG", "8%", "50000", "1000000",
     "Affordable term insurance with high coverage", "Lifetime coverage, income protection, family security", "Yes", "10/20/30 years", "Up to 1 Cr"],
    ["P002", "Niva Bupa Health Insurance", "Health Insurance", "Niva Bupa", "12%", "10000", "500000",
     "Comprehensive health coverage for families", "Cashless treatment, wide network, 24/7 support", "Yes", "Annual renewal", "Up to 50 Lac"],
    ["P003", "POSP Pension Plan", "Pension", "Government", "5%", "100000", "5000000",
     "Secure retirement planning with tax benefits", "Tax-free returns, guaranteed returns, lifetime income", "Yes", "Various terms", "Based on contribution"],
    ["P004", "DSA Investment Plan", "Investment", "Various", "6%", "50000", "2000000",
     "Direct Selling Agent commission structure", "Flexible, market-linked, good returns", "Yes", "Varies", "Market dependent"],
]

for row_idx, product in enumerate(products_data, 2):
    for col_idx, value in enumerate(product, 1):
        ws_products.cell(row=row_idx, column=col_idx, value=value)

set_column_width(ws_products, [80, 150, 120, 120, 100, 100, 100, 200, 200, 80, 150, 150])

# ============================================
# SHEET 7: TASKS
# ============================================
ws_tasks = wb.create_sheet("Tasks")
headers_tasks = ["Task ID", "Task Description", "Assigned To", "Related To", "Related Type",
                 "Status", "Priority", "Due Date", "Created Date", "Completed Date", "Category", "Notes"]
create_header(ws_tasks, 1, headers_tasks, COLOR_TASKS, "000000")

dv_status_task = DataValidation(type="list", formula1='"Pending,In Progress,Completed,On Hold,Cancelled"')
ws_tasks.add_data_validation(dv_status_task)
dv_status_task.add('F2:F1000')

dv_priority = DataValidation(type="list", formula1='"Critical,High,Medium,Low"')
ws_tasks.add_data_validation(dv_priority)
dv_priority.add('G2:G1000')

dv_category = DataValidation(type="list", formula1='"Follow-up,Proposal,Review,Renewal,Documentation"')
ws_tasks.add_data_validation(dv_category)
dv_category.add('K2:K1000')

set_column_width(ws_tasks, [60, 200, 100, 120, 100, 100, 100, 100, 100, 100, 150, 200])

# ============================================
# SHEET 8: COMMUNICATIONS
# ============================================
ws_communications = wb.create_sheet("Communications")
headers_communications = ["Communication ID", "Date & Time", "Contact Name", "Channel", "Type",
                          "Message Subject", "Status", "Outcome", "Next Follow-up",
                          "Duration (min)", "Sent By", "Opened", "Clicked", "Notes"]
create_header(ws_communications, 1, headers_communications, COLOR_COMM)

dv_channel = DataValidation(type="list", formula1='"WhatsApp,Email,Phone Call,In-person,Video Call,SMS,LinkedIn"')
ws_communications.add_data_validation(dv_channel)
dv_channel.add('D2:D1000')

dv_comm_status = DataValidation(type="list", formula1='"Sent,Delivered,Opened,Replied,Scheduled,Failed"')
ws_communications.add_data_validation(dv_comm_status)
dv_comm_status.add('G2:G1000')

set_column_width(ws_communications, [80, 150, 120, 100, 100, 200, 100, 150, 100, 100, 100, 100, 100, 200])

# ============================================
# SHEET 9: LEAD SCORING (Phase C)
# ============================================
ws_scoring = wb.create_sheet("Lead Scoring")
ws_scoring.merge_cells("A1:H1")
ws_scoring["A1"] = "LEAD SCORING ENGINE"
ws_scoring["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws_scoring["A1"].fill = PatternFill(start_color=COLOR_SCORING, end_color=COLOR_SCORING, fill_type="solid")

headers_scoring = ["Lead Name", "Engagement Score", "Firmography Score", "Behavior Score",
                   "Characteristics Score", "Decay Adjustment", "Total Score", "Lead Tier"]
create_header(ws_scoring, 2, headers_scoring, COLOR_SCORING)

scoring_tiers = [
    ["Score Range", "Tier", "Color", "Action"],
    ["80-100", "HOT", "Red", "Call immediately"],
    ["60-79", "WARM", "Orange", "Schedule meeting"],
    ["40-59", "COOL", "Yellow", "Nurture sequence"],
    ["20-39", "COLD", "Blue", "Email campaign"],
    ["0-19", "VERY COLD", "Gray", "Re-engagement"],
]

for row_idx, tier in enumerate(scoring_tiers, 11):
    for col_idx, value in enumerate(tier, 1):
        ws_scoring.cell(row=row_idx, column=col_idx, value=value)

set_column_width(ws_scoring, [150, 150, 150, 150, 150, 150, 120, 150, 150, 150, 150])

# ============================================
# SHEET 10: WORKFLOWS (Phase B)
# ============================================
ws_workflows = wb.create_sheet("Workflows")
ws_workflows.merge_cells("A1:G1")
ws_workflows["A1"] = "AUTOMATION WORKFLOWS"
ws_workflows["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws_workflows["A1"].fill = PatternFill(start_color=COLOR_WORKFLOW, end_color=COLOR_WORKFLOW, fill_type="solid")

headers_workflows = ["Workflow ID", "Workflow Name", "Trigger", "Condition", "Action",
                     "Enabled", "Created Date"]
create_header(ws_workflows, 2, headers_workflows, COLOR_WORKFLOW)

workflows_data = [
    ["W001", "Lead Assignment", "New lead created", "Source = Referral", "Send welcome message", "Yes", datetime.now()],
    ["W002", "Hot Lead Alert", "Lead score > 80", "Status = Qualified", "Notify manager", "Yes", datetime.now()],
    ["W003", "Deal Renewal", "Renewal date today", "Status = Active", "Send renewal reminder", "Yes", datetime.now()],
    ["W004", "Dormant Client", "No contact 30 days", "Status = Active", "Initiate re-engagement", "Yes", datetime.now()],
    ["W005", "Commission Notification", "Commission earned", "Status = Earned", "Send payment alert", "Yes", datetime.now()],
]

for row_idx, workflow in enumerate(workflows_data, 3):
    for col_idx, value in enumerate(workflow, 1):
        ws_workflows.cell(row=row_idx, column=col_idx, value=value)

set_column_width(ws_workflows, [80, 150, 150, 150, 150, 80, 100])

# ============================================
# SHEET 11: EMAIL SEQUENCES (Phase D)
# ============================================
ws_sequences = wb.create_sheet("Email Sequences")
ws_sequences.merge_cells("A1:G1")
ws_sequences["A1"] = "EMAIL DRIP CAMPAIGNS"
ws_sequences["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws_sequences["A1"].fill = PatternFill(start_color=COLOR_SEQUENCES, end_color=COLOR_SEQUENCES, fill_type="solid")

headers_sequences = ["Sequence ID", "Campaign Name", "Email #", "Subject Line", "Send After",
                     "Days", "Enrolled Leads", "Status"]
create_header(ws_sequences, 2, headers_sequences, COLOR_SEQUENCES)

sequences_data = [
    ["S001", "Nurture Sequence", "1", "Welcome to ArthaInvest", "Enrollment", "0", "0", "Active"],
    ["S001", "Nurture Sequence", "2", "Why insurance matters", "Previous email", "3", "0", "Active"],
    ["S001", "Nurture Sequence", "3", "Our solutions for you", "Previous email", "7", "0", "Active"],
    ["S002", "Onboarding Sequence", "1", "Welcome aboard!", "Client creation", "0", "27", "Active"],
    ["S002", "Onboarding Sequence", "2", "Your policy details", "Previous email", "3", "27", "Active"],
]

for row_idx, seq in enumerate(sequences_data, 3):
    for col_idx, value in enumerate(seq, 1):
        ws_sequences.cell(row=row_idx, column=col_idx, value=value)

set_column_width(ws_sequences, [80, 150, 80, 200, 150, 80, 120, 100])

# ============================================
# SHEET 12: COMMUNICATION LOG (Phase E)
# ============================================
ws_comm_log = wb.create_sheet("Communication Hub")
ws_comm_log.merge_cells("A1:G1")
ws_comm_log["A1"] = "UNIFIED COMMUNICATION TRACKING"
ws_comm_log["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws_comm_log["A1"].fill = PatternFill(start_color=COLOR_COMM_TRACK, end_color=COLOR_COMM_TRACK, fill_type="solid")

headers_comm_log = ["Lead/Client Name", "Email Opens", "Email Clicks", "WhatsApp Sent",
                    "WhatsApp Delivered", "Calls Made", "Meetings", "Last Activity", "Next Touchpoint"]
create_header(ws_comm_log, 2, headers_comm_log, COLOR_COMM_TRACK)

comm_data = [
    ["Rajesh Patel", "2", "1", "3", "3", "2", "1", "2026-08-20", "2026-08-25"],
    ["Priya Sharma", "1", "0", "2", "2", "1", "1", "2026-08-18", "2026-08-22"],
]

for row_idx, comm in enumerate(comm_data, 3):
    for col_idx, value in enumerate(comm, 1):
        ws_comm_log.cell(row=row_idx, column=col_idx, value=value)

set_column_width(ws_comm_log, [150, 100, 100, 100, 150, 100, 100, 150, 150])

# ============================================
# SHEET 13: COMMISSIONS
# ============================================
ws_commissions = wb.create_sheet("Commissions")
headers_commissions = ["Commission ID", "Date", "Agent/DSA", "Client Name", "Product",
                       "Commission Amount", "Commission %", "Type", "Status", "Payment Date", "Notes"]
create_header(ws_commissions, 1, headers_commissions, COLOR_COMM_TRACK)

commission_data = [
    ["CM001", datetime.now(), "You", "Priya Sharma", "Tata AIG Term", "40000", "8%", "Initial", "Paid", datetime.now() - timedelta(days=5), "Paid via bank"],
    ["CM002", datetime.now(), "You", "Amit Kumar", "Niva Bupa Health", "96000", "12%", "Trail", "Pending", "", "Waiting for approval"],
]

for row_idx, comm in enumerate(commission_data, 2):
    for col_idx, value in enumerate(comm, 1):
        ws_commissions.cell(row=row_idx, column=col_idx, value=value)

set_column_width(ws_commissions, [80, 100, 120, 120, 120, 120, 100, 100, 100, 100, 200])

# ============================================
# SHEET 14: REPORTS & ANALYTICS
# ============================================
ws_reports = wb.create_sheet("Reports")
ws_reports.merge_cells("A1:D1")
ws_reports["A1"] = "CRM ANALYTICS & REPORTS"
ws_reports["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws_reports["A1"].fill = PatternFill(start_color=COLOR_REPORTS, end_color=COLOR_REPORTS, fill_type="solid")

ws_reports["A3"] = "MONTHLY PERFORMANCE"
ws_reports["A3"].font = Font(bold=True, size=11)

report_headers = ["Metric", "This Month", "Last Month", "YoY Change"]
create_header(ws_reports, 4, report_headers, COLOR_REPORTS)

report_data = [
    ["New Clients Added", "2", "1", "+100%"],
    ["New Leads Generated", "12", "8", "+50%"],
    ["Deals Closed", "3", "2", "+50%"],
    ["Total Commission", "136000", "80000", "+70%"],
    ["Client Retention Rate", "100%", "96%", "+4%"],
    ["Email Open Rate", "28%", "24%", "+4%"],
    ["Conversion Rate", "25%", "20%", "+5%"],
]

for row_idx, row_data in enumerate(report_data, 5):
    for col_idx, value in enumerate(row_data, 1):
        ws_reports.cell(row=row_idx, column=col_idx, value=value)

set_column_width(ws_reports, [200, 120, 120, 120])

# ============================================
# SHEET 15: SETTINGS & CONFIGURATION
# ============================================
ws_settings = wb.create_sheet("Settings")
ws_settings.merge_cells("A1:B1")
ws_settings["A1"] = "CRM CONFIGURATION & SETTINGS"
ws_settings["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws_settings["A1"].fill = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type="solid")

ws_settings["A3"] = "TEAM MEMBERS"
ws_settings["A3"].font = Font(bold=True, size=11)
team_headers = ["Name", "Role", "Email", "Phone"]
create_header(ws_settings, 4, team_headers, COLOR_PRIMARY)

team_data = [
    ["You", "Owner/Agent", "neemailbox555@gmail.com", "Your Mobile"],
]

for row_idx, row_data in enumerate(team_data, 5):
    for col_idx, value in enumerate(row_data, 1):
        ws_settings.cell(row=row_idx, column=col_idx, value=value)

ws_settings["A10"] = "COMMISSION STRUCTURE"
ws_settings["A10"].font = Font(bold=True, size=11)
commission_headers = ["Product", "Initial %", "Trail %", "Bonus %"]
create_header(ws_settings, 11, commission_headers, COLOR_PRIMARY)

commission_structure = [
    ["Tata AIG Term", "8%", "0.5%", "2%"],
    ["Niva Bupa Health", "12%", "1%", "3%"],
    ["POSP Pension", "5%", "0.3%", "1%"],
    ["DSA Investment", "6%", "0%", "2%"],
]

for row_idx, row_data in enumerate(commission_structure, 12):
    for col_idx, value in enumerate(row_data, 1):
        ws_settings.cell(row=row_idx, column=col_idx, value=value)

set_column_width(ws_settings, [200, 120, 120, 200])

# Save the workbook
output_path = r"C:\Users\artha\OneDrive\Desktop\ArthaInvest\ArthaInvest_CRM_ADVANCED.xlsx"
wb.save(output_path)
print("SUCCESS: Advanced CRM created successfully!")
print(f"Location: {output_path}")
print("\nSheets created:")
for idx, sheet in enumerate(wb.sheetnames, 1):
    print(f"  {idx}. {sheet}")
print(f"\nTotal: {len(wb.sheetnames)} complete modules")
print("Status: READY TO USE")
